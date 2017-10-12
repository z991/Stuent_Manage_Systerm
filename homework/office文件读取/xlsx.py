from  openpyxl.reader.excel import load_workbook
#打开文件
wb=load_workbook(filename=r"D:\朱烜宇\day24down\doc\1.xlsx")
print(wb.get_sheet_names()) #每个表格的名称
print(len(wb.get_sheet_names())) #30个表格

sheetnames=wb.get_sheet_names() #list，所有的表格
ws=wb.get_sheet_by_name(sheetnames[0])#第一个表格

print(ws.title) #开头
print(ws.max_row)  #最大的行数
print(ws.max_column) #列数

for  rline  in  range(1,ws.max_row+1):
     for  column in  range(1,ws.max_column+1):
         w1=ws.cell(row=rline,column=  column).value
         print(w1,end=" ")
     print("")