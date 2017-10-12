from  openpyxl.reader.excel import load_workbook
global wl
class readxlsx:
    def __init__(self,path):
        self.path=path
        self.alltext=""
        self.wb=None
    def read(self):
        self.wb = load_workbook(filename=self.path)#载入
        if self.wb !=None:
            sheetnames = self.wb.get_sheet_names()  # list，所有的表格
            for  sheetname in sheetnames:
                ws = self.wb.get_sheet_by_name(sheetname )  #抓取每一个表格
                print(ws.title)  # 开头
                for rline in range(1, ws.max_row + 1):
                    for column in range(1, ws.max_column + 1):
                        w1 = ws.cell(row=rline, column=column).value
                        print(w1, end=" ")
                    print("")



rd=readxlsx(r"D:\朱烜宇\day24down\doc\1.xlsx")
rd.read()
