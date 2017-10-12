from  openpyxl.reader.excel import load_workbook
#打开文件
wb=load_workbook(filename=r"H:\DAY\day24down\doc\1.xlsx")

mylist=[]
sheetnames=wb.get_sheet_names() #list，所有的表格
for  sheetname in sheetnames:
    ws=wb.get_sheet_by_name(sheetname)#第一个表格
    for rline in range(1, ws.max_row + 1):
            for column in range(1, ws.max_column + 1):
                w1= ws.cell(row=rline, column=column).value
                mylist=list(w1)
                print(mylist, end=" ")
            print("")


'''    file = open(r"D:\朱烜宇\day24down\doc\1zxy.txt", "wb")
    if wl != None:
        for mail in wl:
            file.write((mail + "\r\n").encode("utf-8"))
    file.close()'''
